from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from pathlib import Path
from typing import List, Optional, Tuple
from decimal import Decimal, InvalidOperation
import uuid
import time

from models.menu import Menu
from models.menu_item import MenuItem
from models.restaurant import Restaurant

# Base upload directory
UPLOAD_BASE_DIR = Path("uploads/images")

# Expected column headers (with variations) - matches Layout Menu.xlsx format
EXPECTED_HEADERS = [
    "Menu Name",
    "Item Name",
    "Category",
    "Item Description",
    "Price",
    "Image",
    "ID"
]

# Alternative header names (case-insensitive, handles common typos/variations)
HEADER_VARIANTS = {
    "Menu Name": ["menu name", "menuname", "menu", "menu_name", "menuname"],
    "Item Name": ["item name", "itemname", "name", "item_name", "product name", "productname"],
    "Category": ["category", "cat", "categories", "item category", "itemcategory"],
    "Item Description": ["item description", "itemdescription", "description", "item_description", "desc", "details", "item description "],  # Note: trailing space variant
    "Price": ["price", "item price", "itemprice", "item_price", "cost", "amount"],
    "Image": ["image", "item image", "itemimage", "item_image", "picture", "photo", "img"],
    "ID": ["id", "item id", "itemid", "item_id", "external id", "externalid", "external_id", "identifier"]
}


class MenuImportError:
    """Represents an error during menu import"""
    def __init__(self, row: int, item_name: str, error: str):
        self.row = row
        self.item_name = item_name
        self.error = error
    
    def to_dict(self):
        return {
            "row": self.row,
            "item_name": self.item_name,
            "error": self.error
        }


class ImportResult:
    """Result of import operation"""
    def __init__(self):
        self.imported_items = 0
        self.imported_addons = 0
        self.skipped_items = 0
        self.errors: List[MenuImportError] = []
    
    def to_dict(self):
        return {
            "success": len(self.errors) == 0,
            "imported_items": self.imported_items,
            "imported_addons": self.imported_addons,
            "skipped_items": self.skipped_items,
            "errors": [error.to_dict() for error in self.errors]
        }


def normalize_header(header: str) -> str:
    """Normalize header string for comparison"""
    if not header:
        return ""
    return header.strip().lower().replace(" ", "").replace("_", "")


def validate_headers(worksheet) -> Optional[str]:
    """Validate that the Excel file has the correct headers (flexible matching)"""
    if worksheet.max_row < 1:
        return "Excel file is empty"
    
    headers = []
    for col in range(1, min(len(EXPECTED_HEADERS) + 1, worksheet.max_column + 1)):
        cell_value = worksheet.cell(row=1, column=col).value
        headers.append(str(cell_value).strip() if cell_value else "")
    
    # Check if we have at least the required columns
    if len(headers) < len(EXPECTED_HEADERS):
        return f"Expected at least {len(EXPECTED_HEADERS)} columns, found {len(headers)}"
    
    # Check header names (flexible matching with variants)
    for i, expected in enumerate(EXPECTED_HEADERS):
        if i < len(headers):
            header_value = headers[i]
            normalized_header = normalize_header(header_value)
            normalized_expected = normalize_header(expected)
            
            # Check exact match (case-insensitive)
            if normalized_header == normalized_expected:
                continue
            
            # Check against variants
            variants = HEADER_VARIANTS.get(expected, [])
            normalized_variants = [normalize_header(v) for v in variants]
            
            if normalized_header not in normalized_variants:
                # Try fuzzy matching - check if header contains key words
                expected_keywords = expected.lower().split()
                header_lower = header_value.lower()
                
                # If header contains most keywords, consider it valid
                matching_keywords = sum(1 for keyword in expected_keywords if keyword in header_lower)
                if matching_keywords >= len(expected_keywords) * 0.6:  # At least 60% of keywords match
                    continue
                
                # Check if this looks like data instead of a header
                # If column 1 has something that looks like a menu name (not "menu name"), suggest adding headers
                if i == 0 and expected == "Menu Name":
                    # Check if it looks like an actual menu name (has capital letters, multiple words)
                    if any(c.isupper() for c in header_value) and (' ' in header_value or len(header_value) > 5):
                        return (
                            f"Column {i+1} should be '{expected}', but found '{header_value}' which looks like a menu name. "
                            f"Please add a header row with: {', '.join(EXPECTED_HEADERS)}. "
                            f"Your data should start from row 2."
                        )
                
                # If we get here, header doesn't match
                return f"Column {i+1} should be '{expected}' (or similar), found '{header_value}'. Expected columns: {', '.join(EXPECTED_HEADERS)}"
    
    return None


def get_cell_value(worksheet, row: int, col: int) -> Optional[str]:
    """Get cell value as string, returning None if empty"""
    cell = worksheet.cell(row=row, column=col)
    if cell.value is None:
        return None
    return str(cell.value).strip() if str(cell.value).strip() else None


def extract_image_from_cell(image_loader: SheetImageLoader, cell_reference: str, restaurant_id: int, prefix: str = "menu_item") -> Optional[str]:
    """Extract image from Excel cell and save it to file system"""
    try:
        image = image_loader.get(cell_reference)
        if image is None:
            return None
        
        # Create restaurant-specific directory
        restaurant_dir = UPLOAD_BASE_DIR / f"restaurant_{restaurant_id}"
        restaurant_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate unique filename
        timestamp = int(time.time())
        unique_id = str(uuid.uuid4())[:8]
        
        # Determine file extension from image format
        if hasattr(image, 'format'):
            ext_map = {'PNG': '.png', 'JPEG': '.jpg', 'GIF': '.gif', 'WEBP': '.webp'}
            ext = ext_map.get(image.format, '.png')
        else:
            ext = '.png'
        
        filename = f"{prefix}_{timestamp}_{unique_id}{ext}"
        file_path = restaurant_dir / filename
        
        # Save image
        image.save(file_path)
        
        # Return relative path for URL
        return f"/uploads/images/restaurant_{restaurant_id}/{filename}"
    except Exception as e:
        # If image extraction fails, return None (image is optional)
        return None


def parse_price(price_str: Optional[str]) -> Tuple[Optional[Decimal], Optional[str]]:
    """Parse price string to Decimal, returning error message if invalid"""
    if not price_str:
        return None, "Price is required"
    
    try:
        # Remove any currency symbols and whitespace
        cleaned = price_str.replace('$', '').replace(',', '').strip()
        price = Decimal(cleaned)
        if price < 0:
            return None, "Price cannot be negative"
        return price, None
    except (InvalidOperation, ValueError):
        return None, f"Invalid price format: {price_str}"




def detect_if_first_row_is_data(worksheet) -> bool:
    """Detect if first row looks like data instead of headers"""
    if worksheet.max_row < 2:
        return False
    
    # Check if first row values look like data (e.g., menu name, item name, price)
    first_row_values = []
    for col in range(1, min(7, worksheet.max_column + 1)):  # Updated to check up to 6 columns
        cell_value = worksheet.cell(row=1, column=col).value
        first_row_values.append(str(cell_value).strip() if cell_value else "")
    
    # If first row has a price-like value in column 5 (Price column in new format), it's likely data
    if len(first_row_values) >= 5 and first_row_values[4]:
        price_str = first_row_values[4].replace('$', '').replace(',', '').strip()
        try:
            float(price_str)
            return True  # First row has a numeric price, likely data
        except (ValueError, AttributeError):
            pass
    
    # Check if first row values match header patterns
    first_val_lower = first_row_values[0].lower() if first_row_values[0] else ""
    
    # Common header keywords that should be in headers (updated for new format)
    header_keywords = ["menu name", "menuname", "menu", "menu_name", "item name", "itemname", 
                      "category", "cat", "item description", "itemdescription", "description",
                      "price", "item price", "itemprice", "image", "item image", "itemimage"]
    
    # Normalize first value for comparison
    normalized_first = normalize_header(first_row_values[0]) if first_row_values[0] else ""
    
    # If first value doesn't match any header pattern, it might be data
    if first_val_lower and not any(keyword in normalized_first for keyword in [normalize_header(kw) for kw in header_keywords]):
        # Check if it looks like a menu name (has words that could be a menu name)
        # Menu names typically don't start with "menu" or "item" (unless it's "menu name")
        # Common menu names: "Lunch Menu", "Dinner Menu", "Breakfast", etc.
        # If it contains words like "lunch", "dinner", "breakfast", "menu" (as part of menu name), it's likely data
        menu_name_indicators = ["lunch", "dinner", "breakfast", "brunch", "appetizer", "dessert", "drink", "beverage"]
        
        # If first value contains menu name indicators but not header keywords, it's likely data
        if any(indicator in first_val_lower for indicator in menu_name_indicators):
            # Double-check: if it doesn't start with "menu name" or "item", it's probably a menu name
            if not first_val_lower.startswith(("menu name", "menuname", "item")):
                return True  # Likely a menu name in the data row
        
        # Check if second row has proper headers
        if worksheet.max_row >= 2:
            second_row_first = str(worksheet.cell(row=2, column=1).value or "").strip().lower()
            normalized_second = normalize_header(second_row_first)
            if any(keyword in normalized_second for keyword in [normalize_header(kw) for kw in header_keywords]):
                return True  # Headers are in row 2, row 1 is data
            # Also check if row 2 looks like data (has a price in column 5)
            if len(first_row_values) >= 5:
                row2_col5 = str(worksheet.cell(row=2, column=5).value or "").strip()
                if row2_col5:
                    try:
                        float(row2_col5.replace('$', '').replace(',', '').strip())
                        # Both rows have prices, so row 1 is definitely data
                        return True
                    except (ValueError, AttributeError):
                        pass
    
    return False


async def import_menu_from_excel(file_path: str, restaurant: Restaurant) -> ImportResult:
    """
    Import menu items from Excel file
    
    Args:
        file_path: Path to the Excel file
        restaurant: Restaurant object to import items for
    
    Returns:
        ImportResult with statistics and errors
    """
    result = ImportResult()
    
    try:
        # Load workbook
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        
        # First, try to detect if first row is data instead of headers
        first_row_is_data = detect_if_first_row_is_data(worksheet)
        
        # If first row looks like data, skip header validation and start from row 1
        if first_row_is_data:
            start_row = 1
            # Don't add an error, just proceed - headers are optional if data is detected
        else:
            # Validate headers
            start_row = 2
            header_error = validate_headers(worksheet)
            if header_error:
                # Before failing, double-check if row 1 might actually be data
                # Check if row 1 column 1 looks like a menu name (not a header)
                row1_col1 = str(worksheet.cell(row=1, column=1).value or "").strip()
                row1_col1_lower = row1_col1.lower()
                
                # If it contains menu name indicators, treat as data
                menu_indicators = ["lunch", "dinner", "breakfast", "brunch", "appetizer", "dessert"]
                if any(indicator in row1_col1_lower for indicator in menu_indicators):
                    # Row 1 looks like data, proceed without headers
                    start_row = 1
                else:
                    # Check if maybe headers are in row 2 and row 1 is data
                    if worksheet.max_row >= 2:
                        row2_first = str(worksheet.cell(row=2, column=1).value or "").strip().lower()
                        # Check if row 2 has header-like content
                        if any(keyword in row2_first for keyword in ["menu name", "menuname", "menu", "item name", "itemname"]):
                            result.errors.append(MenuImportError(
                                1, 
                                "", 
                                f"Header validation failed: {header_error}. It looks like your headers might be in row 2. "
                                f"Please move headers to row 1. Expected columns: {', '.join(EXPECTED_HEADERS)}"
                            ))
                            return result
                    
                    result.errors.append(MenuImportError(
                        1, 
                        "", 
                        f"Header validation failed: {header_error}. Expected columns in row 1: {', '.join(EXPECTED_HEADERS)}"
                    ))
                    return result
        
        # Initialize image loader
        image_loader = SheetImageLoader(worksheet)
        
        # Process rows (starting from start_row)
        for row_num in range(start_row, worksheet.max_row + 1):
            item_name = None  # Initialize to avoid NameError in exception handler
            try:
                # Get cell values - matches Layout Menu.xlsx format:
                # Column A: Menu Name
                # Column B: Item Name
                # Column C: Category (skipped - not used in model)
                # Column D: Item Description
                # Column E: Price
                # Column F: Image
                # Column G: ID (optional, for duplicate prevention)
                menu_name = get_cell_value(worksheet, row_num, 1)  # Column A
                item_name = get_cell_value(worksheet, row_num, 2)  # Column B
                # Column C (Category) is skipped - not used in MenuItem model
                item_description = get_cell_value(worksheet, row_num, 4)  # Column D
                item_price_str = get_cell_value(worksheet, row_num, 5)  # Column E
                item_image_cell = worksheet.cell(row=row_num, column=6)  # Column F
                external_id = get_cell_value(worksheet, row_num, 7)  # Column G
                
                # Validate required fields
                if not menu_name:
                    result.errors.append(MenuImportError(
                        row_num,
                        item_name or "Unknown",
                        "Menu Name is required"
                    ))
                    continue
                
                if not item_name:
                    result.errors.append(MenuImportError(
                        row_num,
                        "Unknown",
                        "Item Name is required"
                    ))
                    continue
                
                # Get or create menu if it doesn't exist
                menu, created = await Menu.get_or_create(
                    name=menu_name,
                    restaurant=restaurant
                )
                
                # Check for duplicate if external_id is provided
                if external_id:
                    existing_item = await MenuItem.get_or_none(
                        menu=menu,
                        external_id=external_id
                    )
                    if existing_item:
                        result.skipped_items += 1
                        result.errors.append(MenuImportError(
                            row_num,
                            item_name,
                            f"Duplicate item with ID '{external_id}' already exists in menu '{menu_name}'"
                        ))
                        continue
                
                # Parse price
                item_price, price_error = parse_price(item_price_str)
                if price_error:
                    result.errors.append(MenuImportError(
                        row_num,
                        item_name,
                        price_error
                    ))
                    continue
                
                # Extract item image if present (try to extract from cell E)
                item_image_url = None
                cell_ref = item_image_cell.coordinate
                item_image_url = extract_image_from_cell(
                    image_loader, cell_ref, restaurant.id, "menu_item"
                )
                
                # Create menu item
                await MenuItem.create(
                    menu=menu,
                    name=item_name,
                    description=item_description,
                    price=item_price,
                    image_url=item_image_url,
                    external_id=external_id
                )
                result.imported_items += 1
            
            except Exception as e:
                result.errors.append(MenuImportError(
                    row_num,
                    item_name or "Unknown",
                    f"Unexpected error: {str(e)}"
                ))
                continue
        
        workbook.close()
        return result
    
    except Exception as e:
        result.errors.append(MenuImportError(0, "", f"Failed to process Excel file: {str(e)}"))
        return result

